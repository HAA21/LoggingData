import io
import logging
import os
import re
from datetime import datetime
from threading import Lock
from typing import Dict, List, Optional, Tuple

import openpyxl
import requests
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field, field_validator


logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO"),
    format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
)
logger = logging.getLogger("customer-logger")


DATA_PREFIX = "data/"
HEADERS = ["Date", "Customer Name", "Quantity", "Phone Numbers", "Zone"]
CUSTOMERS_FILE = "customers.xlsx"
ZONES_FILE = "zones.xlsx"
IS_VERCEL_RUNTIME = bool(os.getenv("VERCEL"))
LOCAL_DATA_DIR = "/tmp/data" if IS_VERCEL_RUNTIME else os.path.join(os.path.dirname(__file__), "data")


class EntryCreate(BaseModel):
    date: str = Field(..., description="ISO date (YYYY-MM-DD)")
    customer_name: str = Field(..., min_length=1)
    zone: str = Field(..., min_length=1)
    quantity: int = Field(..., ge=0)
    phone_numbers: List[str] = Field(default_factory=list)

    @field_validator("date")
    @classmethod
    def validate_date(cls, value: str) -> str:
        try:
            datetime.strptime(value, "%Y-%m-%d")
        except ValueError as exc:
            raise ValueError("date must be in YYYY-MM-DD format") from exc
        return value

    @field_validator("customer_name", "zone")
    @classmethod
    def validate_required_text(cls, value: str) -> str:
        stripped = value.strip()
        if not stripped:
            raise ValueError("value cannot be empty")
        return stripped

    @field_validator("phone_numbers")
    @classmethod
    def normalize_phone_numbers(cls, value: List[str]) -> List[str]:
        normalized = [v.strip() for v in value if v and v.strip()]
        return normalized


class EntryResponse(BaseModel):
    ok: bool
    filename: str
    month: str
    year: int
    message: str
    storage_mode: str
    warning: Optional[str] = None


class BlobStorage:
    """
    Minimal Vercel Blob REST client using read/write token.
    Falls back to local files when token is not configured.
    """

    def __init__(self) -> None:
        self.token = os.getenv("BLOB_READ_WRITE_TOKEN")
        self.api_base = os.getenv("BLOB_API_BASE_URL", "https://blob.vercel-storage.com")
        self.session = requests.Session()
        # Prefer Blob whenever token exists. This avoids writes to Vercel read-only FS.
        self.local_mode = not bool(self.token)
        if self.local_mode:
            os.makedirs(LOCAL_DATA_DIR, exist_ok=True)
            logger.info("Storage mode: local filesystem (%s)", LOCAL_DATA_DIR)
        else:
            logger.info("Storage mode: Vercel Blob API")

    @property
    def mode(self) -> str:
        return "local" if self.local_mode else "blob"

    def _headers(self) -> Dict[str, str]:
        if not self.token:
            raise RuntimeError("BLOB_READ_WRITE_TOKEN is not configured")
        return {"Authorization": f"Bearer {self.token}"}

    @staticmethod
    def _canonical_monthly_filename(filename: str) -> str:
        """
        Normalize Vercel Blob suffix style:
        april_2026-AbCd123.xlsx -> april_2026.xlsx
        """
        if not filename.lower().endswith(".xlsx"):
            return filename
        match = re.match(r"^(.*?)(-[A-Za-z0-9]+)?(\.xlsx)$", filename, flags=re.IGNORECASE)
        if not match:
            return filename
        return f"{match.group(1)}.xlsx"

    def list_excel_files(self) -> List[str]:
        if self.local_mode:
            names = [name for name in os.listdir(LOCAL_DATA_DIR) if name.lower().endswith(".xlsx")]
            return sorted(names)

        try:
            resp = self.session.get(
                f"{self.api_base}",
                params={"prefix": DATA_PREFIX, "limit": 1000},
                headers=self._headers(),
                timeout=20,
            )
            resp.raise_for_status()
            payload = resp.json()
            blobs = payload.get("blobs", [])
            latest_by_canonical: Dict[str, str] = {}
            modified_by_canonical: Dict[str, str] = {}
            for blob in blobs:
                pathname = blob.get("pathname", "")
                if pathname.lower().endswith(".xlsx") and pathname.startswith(DATA_PREFIX):
                    raw_name = pathname.replace(DATA_PREFIX, "", 1)
                    canonical = self._canonical_monthly_filename(raw_name)
                    modified = blob.get("uploadedAt", "") or blob.get("createdAt", "") or ""
                    # Keep the latest blob per canonical month file.
                    if canonical not in latest_by_canonical or modified > modified_by_canonical.get(canonical, ""):
                        latest_by_canonical[canonical] = raw_name
                        modified_by_canonical[canonical] = modified
            return sorted(latest_by_canonical.keys())
        except Exception:
            logger.exception("Failed to list blob files")
            raise

    def download_file(self, filename: str) -> Optional[bytes]:
        if self.local_mode:
            path = os.path.join(LOCAL_DATA_DIR, filename)
            if not os.path.exists(path):
                return None
            with open(path, "rb") as f:
                return f.read()

        # Vercel Blob may append random suffixes. Resolve by month prefix.
        filename_no_ext = filename[:-5] if filename.lower().endswith(".xlsx") else filename
        lookup_prefix = f"{DATA_PREFIX}{filename_no_ext}"
        try:
            resp = self.session.get(
                f"{self.api_base}",
                params={"prefix": lookup_prefix, "limit": 1000},
                headers=self._headers(),
                timeout=20,
            )
            resp.raise_for_status()
            blobs = resp.json().get("blobs", [])
            if not blobs:
                return None

            # Prefer exact canonical match, otherwise most recently uploaded file.
            target_canonical = self._canonical_monthly_filename(filename)
            candidates = []
            for blob in blobs:
                pathname = blob.get("pathname", "")
                raw_name = pathname.replace(DATA_PREFIX, "", 1) if pathname.startswith(DATA_PREFIX) else pathname
                if self._canonical_monthly_filename(raw_name) == target_canonical:
                    candidates.append(blob)
            if not candidates:
                return None
            candidates.sort(key=lambda b: b.get("uploadedAt", "") or b.get("createdAt", ""), reverse=True)
            blob_url = candidates[0].get("url")
            if not blob_url:
                return None
            file_resp = self.session.get(blob_url, timeout=20)
            file_resp.raise_for_status()
            return file_resp.content
        except Exception:
            logger.exception("Failed to download blob file: %s", filename)
            raise

    def upload_file(self, filename: str, content: bytes) -> None:
        if self.local_mode:
            path = os.path.join(LOCAL_DATA_DIR, filename)
            with open(path, "wb") as f:
                f.write(content)
            return

        pathname = f"{DATA_PREFIX}{filename}"
        try:
            resp = self.session.put(
                f"{self.api_base}/{pathname}",
                params={"addRandomSuffix": "false", "allowOverwrite": "true"},
                headers={**self._headers(), "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
                data=content,
                timeout=30,
            )
            resp.raise_for_status()
        except Exception:
            logger.exception("Failed to upload blob file: %s", filename)
            raise


class ExcelService:
    def __init__(self, storage: BlobStorage) -> None:
        self.storage = storage
        self._customers_cache: List[str] = []
        self._zones_cache: List[str] = []
        self._cache_ready = False
        self._cache_lock = Lock()
        self._write_lock = Lock()

    @staticmethod
    def month_file_from_date(date_text: str) -> Tuple[str, str, int]:
        date_obj = datetime.strptime(date_text, "%Y-%m-%d")
        month_name = date_obj.strftime("%B").lower()
        filename = f"{month_name}_{date_obj.year}.xlsx"
        return filename, month_name, date_obj.year

    def _load_workbook(self, file_bytes: Optional[bytes]) -> openpyxl.Workbook:
        if file_bytes:
            return openpyxl.load_workbook(io.BytesIO(file_bytes))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Entries"
        ws.append(HEADERS)
        return wb

    @staticmethod
    def _to_bytes(workbook: openpyxl.Workbook) -> bytes:
        stream = io.BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    def _extract_customer_names(self, wb: openpyxl.Workbook) -> List[str]:
        ws = wb.active
        names = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            name = row[1]
            if name and str(name).strip():
                names.append(str(name).strip())
        return names

    @staticmethod
    def _extract_single_column_values(wb: openpyxl.Workbook, col_index: int = 1) -> List[str]:
        ws = wb.active
        values = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < col_index:
                continue
            value = row[col_index - 1]
            if value and str(value).strip():
                values.append(str(value).strip())
        return values

    def _ensure_sheet_headers(self, wb: openpyxl.Workbook, headers: List[str]) -> None:
        ws = wb.active
        for idx, header in enumerate(headers, start=1):
            if ws.cell(row=1, column=idx).value != header:
                ws.cell(row=1, column=idx).value = header

    def _read_lookup_values(self, filename: str, header_name: str) -> List[str]:
        file_bytes = self.storage.download_file(filename)
        if not file_bytes:
            return []
        wb = self._load_workbook(file_bytes)
        self._ensure_sheet_headers(wb, [header_name])
        return sorted(set(self._extract_single_column_values(wb, col_index=1)), key=lambda x: x.lower())

    def _upsert_lookup_value(self, filename: str, header_name: str, new_value: str) -> None:
        file_bytes = self.storage.download_file(filename)
        wb = self._load_workbook(file_bytes)
        self._ensure_sheet_headers(wb, [header_name])
        ws = wb.active
        existing = {str(row[0]).strip().lower() for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row and row[0]}
        if new_value.strip().lower() in existing:
            return
        ws.append([new_value.strip()])
        self.storage.upload_file(filename, self._to_bytes(wb))

    def get_all_customers(self) -> List[str]:
        with self._cache_lock:
            if self._cache_ready:
                return self._customers_cache

            try:
                self._customers_cache = self._read_lookup_values(CUSTOMERS_FILE, "Customer Name")
                self._zones_cache = self._read_lookup_values(ZONES_FILE, "Zone")
                self._cache_ready = True
                return self._customers_cache
            except Exception:
                logger.exception("Failed to collect customer names")
                raise

    def get_all_zones(self) -> List[str]:
        with self._cache_lock:
            if self._cache_ready:
                return self._zones_cache
            # primes both caches from lookup files
            self.get_all_customers()
            return self._zones_cache

    def _invalidate_cache(self) -> None:
        with self._cache_lock:
            self._cache_ready = False

    def append_entry(self, entry: EntryCreate) -> Tuple[str, str, int]:
        filename, month_name, year = self.month_file_from_date(entry.date)

        with self._write_lock:
            current = self.storage.download_file(filename)
            wb = self._load_workbook(current)
            ws = wb.active
            self._ensure_sheet_headers(wb, HEADERS)

            ws.append(
                [
                    entry.date,
                    entry.customer_name,
                    int(entry.quantity),
                    ", ".join(entry.phone_numbers),
                    entry.zone,
                ]
            )

            updated = self._to_bytes(wb)
            self.storage.upload_file(filename, updated)
            self._upsert_lookup_value(CUSTOMERS_FILE, "Customer Name", entry.customer_name)
            self._upsert_lookup_value(ZONES_FILE, "Zone", entry.zone)

        self._invalidate_cache()
        return filename, month_name, year

    def get_recent_entries(self, date_text: str, limit: int = 5) -> List[Dict[str, str]]:
        filename, _, _ = self.month_file_from_date(date_text)
        file_bytes = self.storage.download_file(filename)
        if not file_bytes:
            return []

        wb = self._load_workbook(file_bytes)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            rows.append(
                {
                    "date": str(row[0] or ""),
                    "customer_name": str(row[1] or ""),
                    "quantity": str(row[2] if row[2] is not None else ""),
                    "phone_numbers": str(row[3] or ""),
                    "zone": str(row[4] or "") if len(row) > 4 else "",
                }
            )
        return rows[-limit:][::-1]

    def download_month_file(self, month: str, year: int) -> Optional[bytes]:
        filename = f"{month.lower()}_{year}.xlsx"
        return self.storage.download_file(filename)


app = FastAPI(title="Customer Logger", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

storage = BlobStorage()
excel_service = ExcelService(storage)

static_dir = os.path.join(os.path.dirname(__file__), "static")
app.mount("/static", StaticFiles(directory=static_dir), name="static")


@app.get("/", response_class=HTMLResponse)
def root() -> HTMLResponse:
    index_path = os.path.join(static_dir, "index.html")
    if not os.path.exists(index_path):
        raise HTTPException(status_code=500, detail="Frontend file missing")
    return HTMLResponse(open(index_path, "r", encoding="utf-8").read())


@app.get("/api/health")
def health() -> Dict[str, str]:
    warning = ""
    if IS_VERCEL_RUNTIME and storage.local_mode:
        warning = "Running on Vercel without BLOB_READ_WRITE_TOKEN; data is ephemeral."
    return {
        "status": "ok",
        "storage_mode": storage.mode,
        "blob_configured": "true" if bool(storage.token) else "false",
        "warning": warning,
    }


@app.get("/api/customers")
def get_customers() -> Dict[str, List[str]]:
    try:
        return {"customers": excel_service.get_all_customers()}
    except Exception as exc:
        logger.exception("Error in /api/customers")
        raise HTTPException(status_code=500, detail="Could not load customers") from exc


@app.get("/api/zones")
def get_zones() -> Dict[str, List[str]]:
    try:
        return {"zones": excel_service.get_all_zones()}
    except Exception as exc:
        logger.exception("Error in /api/zones")
        raise HTTPException(status_code=500, detail="Could not load zones") from exc


@app.post("/api/entries", response_model=EntryResponse)
def create_entry(entry: EntryCreate) -> EntryResponse:
    try:
        filename, month, year = excel_service.append_entry(entry)
        warning = None
        if IS_VERCEL_RUNTIME and storage.local_mode:
            warning = "Saved to ephemeral storage. Configure BLOB_READ_WRITE_TOKEN to persist data."
        return EntryResponse(
            ok=True,
            filename=filename,
            month=month,
            year=year,
            message="Entry saved successfully",
            storage_mode=storage.mode,
            warning=warning,
        )
    except Exception as exc:
        logger.exception("Error in /api/entries")
        raise HTTPException(status_code=500, detail="Could not save entry") from exc


@app.get("/api/entries/recent")
def get_recent_entries(month: str, year: int) -> Dict[str, List[Dict[str, str]]]:
    try:
        date_text = f"{year}-{datetime.strptime(month, '%B').month:02d}-01"
    except ValueError:
        try:
            month_index = datetime.strptime(month, "%b").month
            date_text = f"{year}-{month_index:02d}-01"
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Month must be full or short English name") from exc

    try:
        entries = excel_service.get_recent_entries(date_text=date_text, limit=5)
        return {"entries": entries}
    except Exception as exc:
        logger.exception("Error in /api/entries/recent")
        raise HTTPException(status_code=500, detail="Could not load recent entries") from exc


@app.get("/api/download/{month}_{year}")
def download_month_file(month: str, year: int):
    try:
        data = excel_service.download_month_file(month=month, year=year)
        if not data:
            raise HTTPException(status_code=404, detail="Monthly file not found")

        filename = f"{month.lower()}_{year}.xlsx"
        return StreamingResponse(
            io.BytesIO(data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Error in /api/download")
        raise HTTPException(status_code=500, detail="Could not download file") from exc


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
